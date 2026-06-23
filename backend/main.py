"""FastAPI backend for the OCR form validator.

Endpoints:
  GET  /records            -> list all stored records (id + holder name)
  GET  /records/{key}      -> fetch a single record by record_no or trailing number
  POST /records            -> create / update a record
  POST /validate           -> OCR an uploaded image and compare to a stored record
"""

import base64
import csv
import io
import json
import re
import sys
import time
from pathlib import Path
from datetime import datetime, timezone

import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel

import segmenter
from comparator import compare_record
from fields import FIELD_ORDER, LABELS, ON_FORM
from ocr_engine import extract_text, extract_text_detail_batch, extract_text_batch

import os

os.environ["FLAGS_use_mkldnn"] = "0"

# scrape_forms.py lives in this backend/ package; add the repo root (one level
# up) to sys.path so it is importable as backend.scrape_forms.
sys.path.insert(0, str(Path(__file__).parent.parent))
import backend.scrape_forms as scrape_forms  # noqa: E402

DATA_DIR = Path(__file__).parent / "data"
# Fallback dataset used until the first scrape produces a scraped_*.json file.
DATA_FILE = DATA_DIR / "all_user_forms_details 2.json"
# Each scrape is written as scraped_<YYYYMMDD_HHMMSS>.json; the validator always
# uses the most recent one, so older scrapes are kept as history.
SCRAPE_GLOB = "scraped_*.json"


def current_data_file() -> Path:
    scraped = sorted(DATA_DIR.glob(SCRAPE_GLOB))
    return scraped[-1] if scraped else DATA_FILE

# The export uses different key names for some fields; map the internal field key
# (used by FIELDS / the comparator) -> the key as it appears in the export.
SOURCE_KEY = {
    "ph_name": "policy_holder_name",
    "ph_address": "policy_holder_address",
    "ph_city": "policy_holder_city",
    "ph_state": "policy_holder_state",
    "ph_zip": "policy_holder_zip",
    "ph_phone": "policy_holder_phone",
    "agent_zip": "agent_zip_code",
    "q1_alcohol": "1.does_the_life_to_be_insured_consume_alcohol/cigarettes/bidis_or_tobacco_in_any_form?",
    "q2_medication": "2._is_the_life_to_be_insured_currently_taking_any_medication_or_drug?",
    "q3a_hypertension": "i)_hypertension/high_blood_pressure",
    "q3b_diabetes": "ii)_diabetes_or_raised_blood_sugar",
    "q3c_cardiovascular": "iii)_cardiovascular_disease,_palpitations,_heart_attack,_stroke,_chest_pain",
    "q3d_genitourinary": "iv)_genitourinary_diseases_e.g._kidney_disorder,_bladder_disorder,_urine_abnormality,_renal_stones_or_genital_organ_disorder",
    "q4_hiv": "4.has_the_life_to_be_insured_ever_been_tested_positive_for_hiv_/_aids,_hepatitis_b_or_c_or_any_sexually_transmitted_disease?",
    "q5_other_insurance": "5.is_the_life_to_be_insured_currently_covered_under_any_health_insurance_policy_with_any_other_company?",
    "q6_involved_pursue": "6.has_the_life_to_be_insured_ever_been_involved_or_is_planning_to_pursue_any?",
    "q7_glasses": "7.does_the_life_to_be_insured_wear_glasses?",
}


def normalize_record(raw: dict) -> dict:
    """Map an exported record's keys to the internal field keys.

    Handles two JSON formats:
    - Old CRM export: uses long source keys like "policy_holder_name",
      "1.does_the_life_to_be_insured...", etc.  SOURCE_KEY maps these.
    - Scraped / already-normalized JSON: already uses the short internal keys
      like "ph_name", "q1_alcohol", etc.  Fall back to the key itself when the
      source-key alias is absent so no values are silently dropped.
    """
    record = {}
    for key in FIELD_ORDER:
        src = SOURCE_KEY.get(key, key)
        if src in raw:
            record[key] = raw[src]
        elif key in raw:          # already-normalized format (scraped JSON)
            record[key] = raw[key]
    return record

app = FastAPI(title="OCR Form Validator")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def load_records() -> list[dict]:
    data_file = current_data_file()
    if not data_file.exists():
        return []
    with data_file.open(encoding="utf-8") as fh:
        raw = json.load(fh)
    return [normalize_record(r) for r in raw]


def save_records(records: list[dict]) -> None:
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    with DATA_FILE.open("w", encoding="utf-8") as fh:
        json.dump(records, fh, indent=2, ensure_ascii=False)


def find_record(records: list[dict], key: str) -> dict | None:
    key = key.strip().lower()
    digits = "".join(ch for ch in key if ch.isdigit())
    for record in records:
        rec_no = str(record.get("record_no", "")).lower()
        if rec_no == key:
            return record
    for record in records:
        rec_no = str(record.get("record_no", "")).lower()
        rec_digits = "".join(ch for ch in rec_no if ch.isdigit())
        if key and key in rec_no:
            return record
        if digits and digits == rec_digits:
            return record
    return None


@app.get("/health")
def health():
    return {"status": "ok"}


def _scrape_timestamp(path: Path) -> str | None:
    m = re.search(r"scraped_(\d{8})_(\d{6})", path.name)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S").isoformat(
            timespec="seconds"
        )
    except ValueError:
        return None


@app.get("/data-status")
def data_status():
    active = current_data_file()
    return {
        "active_file": active.name,
        "is_scraped": active.name.startswith("scraped_"),
        "record_count": len(load_records()),
        "last_scraped_at": _scrape_timestamp(active),
        "scrape_history": sorted(
            (p.name for p in DATA_DIR.glob(SCRAPE_GLOB)), reverse=True
        ),
    }


@app.post("/scrape")
def scrape(
    username: str = Form(default=""),
    password: str = Form(default=""),
):
    """Scrape the CRM and save the result as the active data file.

    Credentials supplied in the request body take priority; if omitted the
    server falls back to ADMIN_USERNAME / ADMIN_PASSWORD from the .env file.
    """
    try:
        records = scrape_forms.scrape_all_forms(
            username=username.strip() or None,
            password=password.strip() or None,
        )
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"Scrape failed: {exc}") from exc
    if not records:
        raise HTTPException(
            status_code=502,
            detail="Scrape returned no records (check credentials / network).",
        )
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = DATA_DIR / f"scraped_{ts}.json"
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as fh:
        json.dump(records, fh, indent=2, ensure_ascii=False)
    return {
        "saved_file": out.name,
        "record_count": len(records),
        "scraped_at": _scrape_timestamp(out),
    }


def parse_excel_records(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty.")

    header_row = rows[0]
    if not header_row:
        raise ValueError("Excel header row is empty.")

    col_map = {}
    normalized_labels = {k.strip().lower(): k for k in LABELS}
    normalized_source_keys = {SOURCE_KEY[k].strip().lower(): k for k in SOURCE_KEY}
    normalized_field_keys = {k.strip().lower(): k for k in FIELD_ORDER}

    for col_idx, cell in enumerate(header_row):
        if cell is None:
            continue
        val = str(cell).strip().lower()

        if val in normalized_field_keys:
            col_map[col_idx] = normalized_field_keys[val]
        elif val in normalized_labels:
            col_map[col_idx] = normalized_labels[val]
        elif val in normalized_source_keys:
            col_map[col_idx] = normalized_source_keys[val]
        else:
            matched_key = None
            for k, label in LABELS.items():
                if label.strip().lower() == val:
                    matched_key = k
                    break
            if matched_key:
                col_map[col_idx] = matched_key
            else:
                for k, src_k in SOURCE_KEY.items():
                    if src_k.strip().lower() == val:
                        matched_key = k
                        break
                if matched_key:
                    col_map[col_idx] = matched_key

    # Check for record_no column mapping
    if not any(k == "record_no" for k in col_map.values()):
        for col_idx, cell in enumerate(header_row):
            if cell is None:
                continue
            val = str(cell).strip().lower()
            if val == "id":
                col_map[col_idx] = "record_no"
                break

    if not any(k == "record_no" for k in col_map.values()):
        raise ValueError("Could not find 'Record No' or 'ID' column in Excel file.")

    records = []
    for row_values in rows[1:]:
        if not any(cell is not None for cell in row_values):
            continue

        record = {}
        for k in FIELD_ORDER:
            record[k] = ""

        for col_idx, val in enumerate(row_values):
            if col_idx in col_map:
                key = col_map[col_idx]
                if val is None:
                    record[key] = ""
                else:
                    record[key] = str(val).strip()

        records.append(record)

    return records


@app.post("/upload-data")
async def upload_data(file: UploadFile = File(...)):
    """Accept a pre-scraped JSON file or Excel file and save it as the active data source.

    It is saved as scraped_<timestamp>.json in DATA_DIR so it becomes the file that
    /records and /validate use going forward.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Invalid file name.")

    filename_lower = file.filename.lower()
    is_json = filename_lower.endswith(".json")
    is_excel = filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls")

    if not is_json and not is_excel:
        raise HTTPException(
            status_code=400, detail="Only .json or .xlsx Excel files are accepted."
        )

    raw = await file.read()

    if is_json:
        try:
            records = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise HTTPException(
                status_code=400, detail=f"Invalid JSON: {exc}"
            ) from exc
    else:
        # Parse Excel using openpyxl
        try:
            records = parse_excel_records(raw)
        except Exception as exc:
            raise HTTPException(
                status_code=400, detail=f"Invalid Excel file or format: {exc}"
            ) from exc

    if not isinstance(records, list):
        raise HTTPException(
            status_code=400,
            detail="Data must represent a list of record objects.",
        )
    if not records:
        raise HTTPException(status_code=400, detail="Uploaded file contains no records.")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = DATA_DIR / f"scraped_{ts}.json"
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as fh:
        json.dump(records, fh, indent=2, ensure_ascii=False)

    return {
        "saved_file": out.name,
        "record_count": len(records),
        "scraped_at": _scrape_timestamp(out),
    }


@app.get("/download-data")
def download_data():
    active = current_data_file()
    if not active.exists():
        raise HTTPException(status_code=404, detail="No data file available")
    return FileResponse(
        active, media_type="application/json", filename=active.name
    )


@app.get("/download-csv")
def download_csv():
    active = current_data_file()
    if not active.exists():
        raise HTTPException(status_code=404, detail="No data file available")
    with active.open(encoding="utf-8") as fh:
        raw = json.load(fh)
    if not raw:
        raise HTTPException(status_code=404, detail="Data file is empty")

    # Known fields first (human-readable label headers), then any extra raw keys.
    known = [(LABELS.get(k, k), SOURCE_KEY.get(k, k)) for k in FIELD_ORDER]
    known_srcs = {src for _, src in known}
    extra: list[str] = []
    seen: set[str] = set()
    for r in raw:
        for k in r:
            if k not in known_srcs and k not in seen:
                seen.add(k)
                extra.append(k)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for label, _ in known] + extra)
    for r in raw:
        writer.writerow(
            [r.get(src, "") for _, src in known] + [r.get(k, "") for k in extra]
        )
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{active.stem}.csv"'
        },
    )


@app.get("/records")
def list_records():
    records = load_records()
    return [
        {
            "record_no": r.get("record_no"),
            "form_no": r.get("form_no"),
            "ph_name": r.get("ph_name"),
        }
        for r in records
    ]


@app.get("/records/{key}")
def get_record(key: str):
    record = find_record(load_records(), key)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    return record


class RecordIn(BaseModel):
    data: dict


@app.post("/records")
def upsert_record(payload: RecordIn):
    record = payload.data
    record_no = str(record.get("record_no", "")).strip()
    if not record_no:
        raise HTTPException(status_code=400, detail="record_no is required")
    records = load_records()
    for idx, existing in enumerate(records):
        if str(existing.get("record_no", "")).strip() == record_no:
            records[idx] = record
            break
    else:
        records.append(record)
    save_records(records)
    return {"saved": record_no}


def match_record_by_number(records: list[dict], ocr_text: str) -> dict | None:
    """Find the record whose record number is printed in the form's header.

    Only the top of the form is searched (where record_no + invoice_no live),
    NOT the whole body. Searching the body matched 5-digit runs that incidentally
    appear inside card / policy / zip numbers, mapping boxes to the wrong record.
    A correct match shows the number twice (record_no and invoice_no), so we
    require it to appear at least twice and ignore weaker single-substring hits."""
    lines = [ln for ln in ocr_text.splitlines() if ln.strip()]
    header = " ".join(lines[:2])  # record_no + invoice_no are on the first line
    ocr_digits = re.sub(r"[^0-9]", "", header)
    if not ocr_digits:
        return None
    best, best_count = None, 1
    for record in records:
        digits = re.sub(r"[^0-9]", "", str(record.get("record_no", "")))
        if not digits:
            continue
        count = ocr_digits.count(digits)
        if count > best_count:
            best, best_count = record, count
    return best


def _recompute_summary(result: dict) -> None:
    """Recompute summary / score / verdict from the (possibly edited) field rows,
    mirroring the comparator so a retry merge stays consistent."""
    fields = result["fields"]
    matched = sum(1 for f in fields if f["status"] == "match")
    mismatched = sum(1 for f in fields if f["status"] == "mismatch")
    checked = matched + mismatched
    result["summary"] = {
        "checked": checked,
        "matched": matched,
        "partial": 0,
        "mismatched": mismatched,
    }
    result["overall_score"] = round(100 * matched / checked, 1) if checked else 0.0
    result["verdict"] = "PASS" if mismatched == 0 else "FAIL"


def _fill_empty_found(result: dict) -> dict:
    """For any on-form field (serials 2-54, e.g. Blood Group, Nominee State) where
    OCR read nothing, fall back to the CRM value so an unreadable mark doesn't
    fail the form. Computed/CRM-only fields (Total Amount, etc.) are untouched.

    This handles two cases:
    - status == "mismatch" with empty found: OCR alignment placed an empty span
    - status == "match" with empty found: field had no CRM value (already handled
      upstream, but guard here too)
    """
    changed = False
    for f in result["fields"]:
        if not ON_FORM.get(f["field"]):
            continue
        # Don't promote a genuine ocr_missed to a match — that would hide the
        # fact that the aligner found nothing for this field.
        if f["status"] == "ocr_missed":
            continue
        found_val = str(f.get("found", "")).strip()
        expected_val = str(f.get("expected", "")).strip()
        if not found_val and expected_val and f["status"] in ("mismatch",):
            f["found"] = f["expected"]
            f["status"] = "match"
            f["score"] = 100.0
            f["box"] = None
            changed = True
    if changed:
        _recompute_summary(result)
    return result


def _fill_empty_found_excel(result: dict) -> dict:
    """More aggressive fallback for the OCR-to-Excel export path.

    Unlike _fill_empty_found (used for validation), this fills ANY on-form
    field where OCR alignment returned an empty 'found' value, regardless of
    status. This prevents the Excel output from having blank cells for fields
    whose values exist in the CRM but were missed by the positional aligner
    (e.g. Policy Holder address/city/state/zip, and all Yes/No questions when
    many consecutive identical answers confuse the DP).
    """
    changed = False
    for f in result["fields"]:
        if not ON_FORM.get(f["field"]):
            continue
        found_val = str(f.get("found", "")).strip()
        expected_val = str(f.get("expected", "")).strip()
        if not found_val and expected_val:
            f["found"] = f["expected"]
            # Keep the original status so the caller knows this was a fallback;
            # for Excel we only care about the value, not the status.
            changed = True
    if changed:
        _recompute_summary(result)
    return result


def _encode_crop(crop: np.ndarray, boxes: list) -> str:
    """JPEG-encode a crop as a base64 data URL, drawing a red box around each
    given (x0,y0,x1,y1) region (used to flag the mismatched values)."""
    img = crop.copy()
    h, w = img.shape[:2]
    pad = max(2, round(min(h, w) * 0.01))
    for box in boxes:
        if not box:
            continue
        x0, y0, x1, y1 = (int(round(v)) for v in box)
        cv2.rectangle(
            img,
            (max(0, x0 - pad), max(0, y0 - pad)),
            (min(w - 1, x1 + pad), min(h - 1, y1 + pad)),
            (0, 0, 255),
            max(2, round(min(h, w) * 0.004)),
        )
    buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 75])[1]
    return "data:image/jpeg;base64," + base64.b64encode(buf).decode("ascii")


def _group_forms(ocr_details: list, meta: list) -> list:
    """Group crops into whole forms, rejoining a record split across pages: a
    top-partial (record marker but clipped/incomplete) waits for its bottom
    continuation in the SAME column; complete forms stand alone. Returns lists of
    crop indices (each list = one form, in top->bottom order)."""
    from image_to_excel import _is_form_start, _is_complete

    pending: list = []  # [[member_indices], col]
    groups: list = []
    for i, d in enumerate(ocr_details):
        toks = d["text"].split()
        col = meta[i].get("col", 0.5)
        if not toks:
            groups.append([i])
        elif _is_form_start(toks):
            if _is_complete(toks):
                groups.append([i])
            else:
                pending.append([[i], col])
        elif pending:  # bottom-half -> nearest open top-partial by column
            j = min(range(len(pending)), key=lambda k: abs(pending[k][1] - col))
            members = pending.pop(j)[0]
            members.append(i)
            groups.append(members)
        else:
            groups.append([i])  # orphan bottom-half -> its own (will be unmatched)
    for members, _ in pending:  # top-partials that never got a continuation
        groups.append(members)
    return groups


def _validate_form_group(members, ocr_details, crops_bgr, meta, records):
    """Validate one form (one or more crops). For a multi-crop (split) form the
    OCR text is merged before matching/comparison, and each crop is returned in
    `sources` annotated only with the field boxes that fall on it. Returns
    (home_image_name, result)."""
    merged_text = " ".join(ocr_details[m]["text"] for m in members)
    merged_boxes: list = []
    offsets: list = []
    running = 0
    for m in members:
        offsets.append(running)
        merged_boxes.extend(ocr_details[m]["token_boxes"])
        running += len(ocr_details[m]["text"].split())

    def member_of(tok_start):
        if tok_start is None:
            return 0
        k = 0
        for idx, off in enumerate(offsets):
            if tok_start >= off:
                k = idx
        return k

    src_meta = [
        {"image_name": meta[m]["image_name"], "box": meta[m]["box"]} for m in members
    ]
    home = meta[members[-1]]["image_name"]  # the crop that completed it (this batch)

    record = match_record_by_number(records, merged_text)
    if not record:
        sources = [
            {**src_meta[k], "image": _encode_crop(crops_bgr[members[k]], [])}
            for k in range(len(members))
        ]
        return home, {
            "image_name": home,
            "box": meta[members[-1]]["box"],
            "sources": sources,
            "image": sources[0]["image"],
            "matched": False,
            "record_no": None,
            "ocr_text": merged_text,
            "message": "No matching record found for this form",
        }

    result = compare_record(record, merged_text, merged_boxes)
    per_member: list = [[] for _ in members]
    for f in result["fields"]:
        if f.get("box"):
            per_member[member_of(f.get("tok_start"))].append(f["box"])
    sources = [
        {**src_meta[k], "image": _encode_crop(crops_bgr[members[k]], per_member[k])}
        for k in range(len(members))
    ]
    result.update(
        {
            "image_name": home,
            "box": meta[members[-1]]["box"],
            "sources": sources,
            "image": sources[0]["image"],
            "matched": True,
            "record_no": record.get("record_no"),
            "form_no": record.get("form_no"),
            "ph_name": record.get("ph_name"),
            "ocr_text": merged_text,
        }
    )
    return home, result


@app.post("/validate-batch")
async def validate_batch(images: list[UploadFile] = File(...)):
    started = time.time()
    records = load_records()

    # Detect & crop every form box across ALL uploaded images first, then OCR the
    # whole pile in one parallel pass (one box = one OCR job). `meta` keeps each
    # crop tied back to its source image, position and column (for stitching).
    encoded: list[bytes] = []
    crops_bgr: list[np.ndarray] = []
    meta: list[dict] = []
    image_errors: dict[str, str] = {}
    order: list[str] = []
    for up in images:
        name = up.filename or f"image_{len(order)}"
        order.append(name)
        raw = await up.read()
        if not raw:
            image_errors[name] = "Empty image"
            continue
        bgr = cv2.imdecode(np.frombuffer(raw, dtype=np.uint8), cv2.IMREAD_COLOR)
        if bgr is None:
            image_errors[name] = "Could not decode image"
            continue
        W = bgr.shape[1] or 1
        boxes = segmenter.detect_form_boxes(bgr)
        crops = segmenter.crop_boxes(bgr, boxes)
        for box_i, (crop, (x, y, w, h)) in enumerate(zip(crops, boxes)):
            encoded.append(cv2.imencode(".png", crop)[1].tobytes())
            crops_bgr.append(crop)
            meta.append({"image_name": name, "box": box_i, "col": (x + w / 2) / W})

    ocr_details = extract_text_detail_batch(encoded)

    # Rejoin records split across pages, then validate each whole form. A merged
    # form is filed under the image of the crop that completed it (this batch).
    by_image: dict[str, list] = {name: [] for name in order}
    for members in _group_forms(ocr_details, meta):
        home, result = _validate_form_group(
            members, ocr_details, crops_bgr, meta, records
        )
        by_image.setdefault(home, []).append(result)

    images_out = [
        {
            "image_name": name,
            "error": image_errors.get(name),
            "forms_detected": len(by_image.get(name, [])),
            "results": by_image.get(name, []),
        }
        for name in order
    ]

    return {
        "image_count": len(order),
        "forms_detected": sum(len(v) for v in by_image.values()),
        "images": images_out,
        "elapsed_seconds": round(time.time() - started, 1),
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }


@app.post("/validate")
async def validate(form_no: str = Form(...), image: UploadFile = File(...)):
    record = find_record(load_records(), form_no)
    if not record:
        raise HTTPException(
            status_code=404,
            detail=f"No stored record found for '{form_no}'",
        )

    started = time.time()
    image_bytes = await image.read()
    if not image_bytes:
        raise HTTPException(status_code=400, detail="Empty image upload")

    try:
        ocr_text = extract_text(image_bytes)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"OCR failed: {exc}") from exc

    result = compare_record(record, ocr_text)  # validation
    # No fill_empty_found: ocr_missed and mismatch must stay visible.
    result.update(
        {
            "record_no": record.get("record_no"),
            "form_no": record.get("form_no"),
            "ph_name": record.get("ph_name"),
            "image_name": image.filename,
            "ocr_text": ocr_text,
            "elapsed_seconds": round(time.time() - started, 1),
            "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        }
    )
    return result


from image_to_excel import (
    HEADER as EXTRACT_HEADER,
    extract_row,
    stitch_texts_by_column,
)


@app.post("/extract")
async def extract_forms(
    images: list[UploadFile] = File(default=[]),
    carry: str = Form(default="[]"),
    flush: str = Form(default="false")
):
    started = time.time()
    try:
        pending = json.loads(carry)
    except Exception:
        pending = []

    flush_bool = str(flush).lower() == "true"

    if flush_bool or not images:
        rows = [extract_row(p[0] if not isinstance(p, str) else p) for p in pending]
        return {
            "header": EXTRACT_HEADER,
            "rows": rows,
            "count": len(rows),
            "pending": [],
            "images": [],
            "elapsed_seconds": round(time.time() - started, 1),
            "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds")
        }

    encoded: list[bytes] = []
    cols: list[float] = []  # each crop's x-center as a fraction of page width
    order: list[str] = []
    image_crop_counts: list[int] = []

    for up in images:
        name = up.filename or f"image_{len(order)}"
        order.append(name)
        raw = await up.read()
        if not raw:
            image_crop_counts.append(0)
            continue
        bgr = cv2.imdecode(np.frombuffer(raw, dtype=np.uint8), cv2.IMREAD_COLOR)
        if bgr is None:
            image_crop_counts.append(0)
            continue

        W = bgr.shape[1] or 1
        boxes = segmenter.detect_form_boxes(bgr)
        crops = segmenter.crop_boxes(bgr, boxes)
        image_crop_counts.append(len(crops))
        for crop, (x, y, w, h) in zip(crops, boxes):
            encoded.append(cv2.imencode(".png", crop)[1].tobytes())
            cols.append((x + w / 2) / W)

    texts = extract_text_batch(encoded) if encoded else []

    completed = []
    new_pending = pending
    idx = 0
    for count in image_crop_counts:
        items = list(zip(texts[idx : idx + count], cols[idx : idx + count]))
        c, new_pending = stitch_texts_by_column(new_pending, items)
        completed.extend(c)
        idx += count

    rows = [extract_row(t) for t in completed]
    
    return {
        "header": EXTRACT_HEADER,
        "rows": rows,
        "count": len(rows),
        "pending": new_pending,
        "images": order,
        "elapsed_seconds": round(time.time() - started, 1),
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds")
    }
